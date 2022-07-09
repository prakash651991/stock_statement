import pandas as pd
import numpy as np
from sqlalchemy import create_engine
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
pd.options.mode.chained_assignment = None
from urllib.parse import quote


# Log.txt:
with open('log.txt', 'a') as f:
    f.writelines(f'-------------2. Data Processing Started----------------\n')

raw_data = pd.read_excel(r'raw_data.xlsx')
raw_data['OD_Days'] = raw_data['OD_Days'].astype(int)

# Declare Variable from Excel
rule_sheet = pd.read_excel(r'rule_sheet.xlsx')
Lendor_tag_todolist = rule_sheet['Lendor_tag'].values.tolist()
Required_Encumbrance_todolist = rule_sheet['Required_Encumbrance'].values.tolist()
DPD_allowed_todolist = rule_sheet['DPD_allowed'].values.tolist()
Disbursement_Date_todolist = rule_sheet['Disbursement_Date'].values.tolist()
Product_todolist = rule_sheet['Product'].values.tolist()

#change data type
Disbursement_Date_todolist = pd.to_datetime(Disbursement_Date_todolist, format='%Y-%m-%d')

for i in range(len(Lendor_tag_todolist)):
#for i in range(1):
    Lendor_tag = Lendor_tag_todolist[i]
    Required_Encumbrance = Required_Encumbrance_todolist[i]
    DPD_allowed = DPD_allowed_todolist[i]
    Disbursement_Date = Disbursement_Date_todolist[i]
    Product_l = Product_todolist[i].split(',')
    #print(type(Product_l))
    chk = Product_l[0]

    cbs_single_lendor_data = raw_data[raw_data['funding_txn_remark'] == Lendor_tag].copy()
    # apply filters
    cbs_single_lendor_data['status'] = np.where(~cbs_single_lendor_data['account_status'].str.match('Open'), 0, 1)
    cbs_single_lendor_data['status'] = np.where(cbs_single_lendor_data['OD_Days'] > DPD_allowed, 0,
                                                cbs_single_lendor_data['status'])

    #cbs_single_lendor_data.to_excel('before.xlsx')
    if chk != "'ALL'":
        # product filter
        cleaned_product = []
        for product in Product_l:
            prodstring = product.split("'")
            cleaned_product.append(prodstring[1])
            # print(cleaned_product)
        # print(cleaned_product)
        cbs_single_lendor_data['status'] = np.where(~cbs_single_lendor_data['Product'].isin(cleaned_product), 0,
                                                    cbs_single_lendor_data['status'])
    else:

        #cbs_single_lendor_data.to_excel('after.xlsx')


        # seperate selection and removed accounts
        cbs_single_lendor_selected = cbs_single_lendor_data[cbs_single_lendor_data['status'] == 1]
        cbs_single_lendor_removed = cbs_single_lendor_data[cbs_single_lendor_data['status'] == 0]

        cbs_single_lendor_selected.reset_index(drop=True, inplace=True)

        # remove column
        cbs_single_lendor_selected = cbs_single_lendor_selected.drop(['status'], axis=1)
        cbs_single_lendor_removed = cbs_single_lendor_removed.drop(['status'], axis=1)
        #add column tag
        cbs_single_lendor_removed['tag'] = Lendor_tag
        cbs_single_lendor_selected['tag'] = Lendor_tag

        cnx = create_engine('mysql+pymysql://analytics:%s@34.93.197.76:61306/analytics' % quote('Secure@123'))
        cbs_single_lendor_removed.to_sql(con=cnx, name='stock_report_removed_acc', if_exists='append', index=False)
        cnx.dispose()

        cbs_single_lendor_selected_pos_Value = cbs_single_lendor_selected['POS'].sum()

        print(f"selected {Lendor_tag} posvalue:{cbs_single_lendor_selected_pos_Value}")

if cbs_single_lendor_selected_pos_Value > Required_Encumbrance:
    cbs_single_lendor_selected['cum_sum'] = cbs_single_lendor_selected['POS'].cumsum()
    last_index_num = np.searchsorted(cbs_single_lendor_selected['cum_sum'], Required_Encumbrance)
    # select & removed accounts
    excess_accounts_to_removed = cbs_single_lendor_selected.loc[last_index_num + 1:len(cbs_single_lendor_selected)]
    final_single_lendor_selected = cbs_single_lendor_selected.loc[0:last_index_num]

    cbs_single_lendor_selected.reset_index(drop=True, inplace=True)

    # remove column
    final_single_lendor_selected = final_single_lendor_selected.drop(['cum_sum'], axis=1)
    excess_accounts_to_removed = excess_accounts_to_removed.drop(['cum_sum'], axis=1)

    # cbs_single_lendor_selected.to_excel(f"{Lendor_tag}_selected_output.xlsx")

    0

    # Log.txt
    with open('log.txt', 'a') as f:
        f.writelines(f'{Lendor_tag} is processed.\n')

    excess_accounts_to_removed.to_sql(con=cnx, name='stock_report_removed_excess_acc', if_exists='append', index=False)
    cnx.dispose()

    # excess accounts to excel
    excess_accounts_to_removed = excess_accounts_to_removed.drop(['tag'], axis=1)
    wb = load_workbook(filename="Unencumbered_data.xlsx")
    ws = wb["Sheet1"]
    for r in dataframe_to_rows(excess_accounts_to_removed, index=False,
                               header=False):  # No index and don't append the column headers
        ws.append(r)
    wb.save("Unencumbered_data.xlsx")

elif cbs_single_lendor_selected_pos_Value < Required_Encumbrance:

    Req_Encumbrance = Required_Encumbrance - cbs_single_lendor_selected_pos_Value

    cnx = create_engine('mysql+pymysql://analytics:%s@34.93.197.76:61306/analytics' % quote('Secure@123'))
    cbs_single_lendor_selected.to_sql(con=cnx, name='stock_report_selected_acc_set2', if_exists='append', index=False)
    cnx.dispose()

    fresh_accounts = pd.read_excel(r'Unencumbered_data.xlsx')
    fresh_accounts['OD_Days'] = fresh_accounts['OD_Days'].astype(int)

    # filter selection
    fresh_accounts['status'] = np.where(fresh_accounts['DisbursementDate'] > Disbursement_Date, 1, 0)
    # fresh_accounts.to_excel("step1_disb_filter.xlsx")
    fresh_accounts['status'] = np.where(fresh_accounts['OD_Days'] > DPD_allowed, 0, fresh_accounts['status'])

    fresh_accounts.to_excel('before.xlsx')

    if chk != "'ALL'":
        # product filter
        cleaned_product = []
        for product in Product_l:
            prodstring = product.split("'")
            cleaned_product.append(prodstring[1])
        print(cleaned_product)
        fresh_accounts['status'] = np.where(~fresh_accounts['Product'].isin(cleaned_product), 0,
                                            fresh_accounts['status'])
    else:

        # product_filter

        # fresh_accounts.to_excel('after.xlsx')

        # Sort by selection
        fresh_accounts.sort_values(by=['status', 'DisbursementDate'], inplace=True, ascending=[False, True])
        fresh_accounts.reset_index(drop=True, inplace=True)
        # fresh_accounts.to_excel('after_sort.xlsx')

        # fresh_accounts.to_excel(f"{Lendor_tag}step3_sort.xlsx")
        ####fresh_Unencumbered_data.reset_index(drop=True, inplace=True)

        # calculate cumulative sum
        fresh_accounts['cum_sum'] = fresh_accounts['POS'].cumsum()
        lst_index_num = np.searchsorted(fresh_accounts['cum_sum'], Req_Encumbrance)

        fresh_accounts.loc[0:lst_index_num, 'status'] = 2
        fresh_accounts.reset_index(drop=True, inplace=True)
        #
        Selected_additional_Accounts = fresh_accounts[fresh_accounts['status'] == 2]
        fresh_accounts = fresh_accounts[fresh_accounts['status'] == 1]
        Selected_additional_Accounts.reset_index(drop=True, inplace=True)
        Selected_additional_Accounts['tag'] = Lendor_tag
        Selected_additional_Accounts.drop(['cum_sum'], axis=1, inplace=True)
        Selected_additional_Accounts.drop(['status'], axis=1, inplace=True)

        #
        cnx = create_engine('mysql+pymysql://analytics:%s@34.93.197.76:61306/analytics' % quote('Secure@123'))
        Selected_additional_Accounts.to_sql(con=cnx, name='stock_report_selected_acc_set2', if_exists='append',
                                            index=False)
        cnx.dispose()
        print(f'{Lendor_tag}  - stock_report_selected_acc_set2 - is updated')

        # Log.txt
        with open('log.txt', 'a') as f:
            f.writelines(f'{Lendor_tag} is processed.\n')

        fresh_accounts.reset_index(drop=True, inplace=True)
        # export fresh Unencumbered_data
        fresh_accounts.to_excel('Unencumbered_data.xlsx', index=False)

    if cbs_single_lendor_selected_pos_Value > Required_Encumbrance:
        cbs_single_lendor_selected['cum_sum'] = cbs_single_lendor_selected['POS'].cumsum()
        last_index_num = np.searchsorted(cbs_single_lendor_selected['cum_sum'], Required_Encumbrance)
        # select & removed accounts
        excess_accounts_to_removed = cbs_single_lendor_selected.loc[last_index_num + 1:len(cbs_single_lendor_selected)]
        final_single_lendor_selected = cbs_single_lendor_selected.loc[0:last_index_num]

        cbs_single_lendor_selected.reset_index(drop=True, inplace=True)

        # remove column
        final_single_lendor_selected = final_single_lendor_selected.drop(['cum_sum'], axis=1)
        excess_accounts_to_removed = excess_accounts_to_removed.drop(['cum_sum'], axis=1)

        # cbs_single_lendor_selected.to_excel(f"{Lendor_tag}_selected_output.xlsx")

        cnx = create_engine('mysql+pymysql://analytics:%s@34.93.197.76:61306/analytics' % quote('Secure@123'))
        final_single_lendor_selected.to_sql(con=cnx, name='stock_report_selected_acc', if_exists='append', index=False)
        print(f'{Lendor_tag}  - stock_report_selected_acc set 1 - is updated')

        # Log.txt
        with open('log.txt', 'a') as f:
            f.writelines(f'{Lendor_tag} is processed.\n')

        excess_accounts_to_removed.to_sql(con=cnx, name='stock_report_removed_excess_acc', if_exists='append',
                                          index=False)
        cnx.dispose()

        # excess accounts to excel
        excess_accounts_to_removed = excess_accounts_to_removed.drop(['tag'], axis=1)
        wb = load_workbook(filename="Unencumbered_data.xlsx")
        ws = wb["Sheet1"]
        for r in dataframe_to_rows(excess_accounts_to_removed, index=False,
                                   header=False):  # No index and don't append the column headers
            ws.append(r)
        wb.save("Unencumbered_data.xlsx")

    elif cbs_single_lendor_selected_pos_Value < Required_Encumbrance:

        Req_Encumbrance = Required_Encumbrance - cbs_single_lendor_selected_pos_Value

        cnx = create_engine('mysql+pymysql://analytics:%s@34.93.197.76:61306/analytics' % quote('Secure@123'))
        cbs_single_lendor_selected.to_sql(con=cnx, name='stock_report_selected_acc_set2', if_exists='append',
                                          index=False)
        cnx.dispose()

        fresh_accounts = pd.read_excel(r'Unencumbered_data.xlsx')
        fresh_accounts['OD_Days'] = fresh_accounts['OD_Days'].astype(int)

        # filter selection
        fresh_accounts['status'] = np.where(fresh_accounts['DisbursementDate'] > Disbursement_Date, 1, 0)
        # fresh_accounts.to_excel("step1_disb_filter.xlsx")
        fresh_accounts['status'] = np.where(fresh_accounts['OD_Days'] > DPD_allowed, 0, fresh_accounts['status'])

        fresh_accounts.to_excel('before.xlsx')

        if chk != "'ALL'":
            # product filter
            cleaned_product = []
            for product in Product_l:
                prodstring = product.split("'")
                cleaned_product.append(prodstring[1])
            print(cleaned_product)
            fresh_accounts['status'] = np.where(~fresh_accounts['Product'].isin(cleaned_product), 0,
                                                fresh_accounts['status'])
        else:

            # product_filter

            # fresh_accounts.to_excel('after.xlsx')

            # Sort by selection
            fresh_accounts.sort_values(by=['status', 'DisbursementDate'], inplace=True, ascending=[False, True])
            fresh_accounts.reset_index(drop=True, inplace=True)
            # fresh_accounts.to_excel('after_sort.xlsx')

            # fresh_accounts.to_excel(f"{Lendor_tag}step3_sort.xlsx")
            ####fresh_Unencumbered_data.reset_index(drop=True, inplace=True)

            # calculate cumulative sum
            fresh_accounts['cum_sum'] = fresh_accounts['POS'].cumsum()
            lst_index_num = np.searchsorted(fresh_accounts['cum_sum'], Req_Encumbrance)

            fresh_accounts.loc[0:lst_index_num, 'status'] = 2
            fresh_accounts.reset_index(drop=True, inplace=True)
            #
            Selected_additional_Accounts = fresh_accounts[fresh_accounts['status'] == 2]
            fresh_accounts = fresh_accounts[fresh_accounts['status'] == 1]
            Selected_additional_Accounts.reset_index(drop=True, inplace=True)
            Selected_additional_Accounts['tag'] = Lendor_tag
            Selected_additional_Accounts.drop(['cum_sum'], axis=1, inplace=True)
            Selected_additional_Accounts.drop(['status'], axis=1, inplace=True)

            #
            cnx = create_engine('mysql+pymysql://analytics:%s@34.93.197.76:61306/analytics' % quote('Secure@123'))
            Selected_additional_Accounts.to_sql(con=cnx, name='stock_report_selected_acc_set2', if_exists='append',
                                                index=False)
            cnx.dispose()
            print(f'{Lendor_tag}  - stock_report_selected_acc_set2 - is updated')

            # Log.txt
            with open('log.txt', 'a') as f:
                f.writelines(f'{Lendor_tag} is processed.\n')

            fresh_accounts.reset_index(drop=True, inplace=True)
            # export fresh Unencumbered_data
            fresh_accounts.to_excel('Unencumbered_data.xlsx', index=False)

# Log.txt:
with open('log.txt', 'a') as f:
    f.writelines(f'-------------2. Data Processing completed----------------\n')





