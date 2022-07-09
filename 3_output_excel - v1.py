from openpyxl import load_workbook
from sqlalchemy import create_engine
import pandas as pd
import datetime
from urllib.parse import quote
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
pd.options.mode.chained_assignment = None


#log.txt
with open('log.txt', 'a') as f:
    f.writelines(f'-------------3. Export to Excel started----------------\n')

cnx = create_engine('mysql+pymysql://analytics:%s@34.93.197.76:61306/analytics' % quote('Secure@123'))
final_sel_out_query= f"call analytics.stock_statement_final_selected_out"
final_sel_output = pd.read_sql(final_sel_out_query, cnx, index_col=None)
cnx.dispose()
#change datatype urn
final_sel_output['URN'] = final_sel_output['URN'].values.astype(str)

unique_tag = final_sel_output["tag"].unique()
#print(unique_tag)

#rule sheet for summary data
summ_out = pd.read_excel(r'rule_sheet.xlsx')

for i in unique_tag:
    #data sheet
    new_df = final_sel_output[final_sel_output["tag"]==i]
    new_df = new_df[
        ['KGFS', 'branch', 'centre_name', 'funder_name', 'funding_txn_type', 'funding_txn_remark', 'URN',
         'AccountNumber', 'Product', 'DisbursementDate', 'SanctionedAmount', 'disb_amount', 'Interest_Rate',
         'Installment_Amount', 'Repayment_Frequency', 'MaturityDate', 'POS', 'OD_Days', 'loan_purpose',
         'loan_purpose_detail', 'Age', 'gender', 'customer_name', 'father_name', 'spouse_name', 'id_proof',
         'id_proof_no', 'address_proof', 'address_proof_no', 'mobile_number', 'address', 'district', 'state']]

    wb = load_workbook(r'stock_statement_template.xlsx')
    ws = wb.worksheets[1]


    for r in dataframe_to_rows(new_df, index=False, header=True):
        ws.append(r)

    #summery sheet
    summ_output = pd.read_excel(r'rule_sheet.xlsx')
    summary_df = summ_output[summ_output["Lendor_tag"] == i]
    summary_df.reset_index(drop=True, inplace=True)
    result_1 = summary_df.at[0, 'Debt_Outstanding_from_Tally']
    result_2 = summary_df.at[0, 'Book_Debts']
    ws = wb.worksheets[0]

    ws.cell(row=6, column=4).value = result_1
    ws.cell(row=7, column=4).value = result_2
    wb.save(f'{i}.xlsx')
    # log.txt
    with open('log.txt', 'a') as f:
        f.writelines(f'Stock Statement Report - "{i}.xlsx" is created\n')

#log.txt
with open('log.txt', 'a') as f:
    f.writelines(f'-------------3. Export to Excel is Completed----------------\n')